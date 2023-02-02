import unittest
import os
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from decimal import Decimal
import math
from datetime import datetime
import time
import json
import chromedriver_autoinstaller
import time
import zipfile
import json
import shutil
import os
import pywinauto.keyboard
import pywinauto.mouse
import csv
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.chart.axis import ChartLines

class opAMP1(unittest.TestCase):

    def setUp(self):
        # driver instance
        chromedriver_autoinstaller.install()
        self.driver = webdriver.Chrome()
        with open(r'paths.json') as d:
            self.nimbleData = json.load(d)['Nimble'][0]

    def test_export(self):
        driver = self.driver
        driver.set_window_position(-1000, 0)
        driver.maximize_window()
        driver.get('https://beta-tools.analog.com/noise/')
        if (self.nimbleData['filter_frequency'] != 0): #self.nimbleData['resistance_input'
            driver.get('https://beta-tools.analog.com/noise/#session=HO1dbVi7oU273IbstN6GiA&step=nsONZTDZTletQ0rqfFMQ0g')
            print("Filter was added.")
        else:
            driver.get('https://beta-tools.analog.com/noise/#session=0460LuqFaUGaPwEazOdZ9w&step=ZiZ21WxLTiiX7jikxx07Aw')
            print("No filther was added.")
        time.sleep(1)

        #converting kino, Mega
        d1 = {'k': 1000,'M': 1000000}
        def text_to_num1(text):
            if text[-1] in d1:
                num, magnitude = text[:-1], text[-1]
                return float(num) * d1[magnitude]
            else:
                return Decimal(text)
        new_rvalue = text_to_num1(self.nimbleData['rvalue'])

        #converting from femto, pico, nano, micro
        d2 = {'f': 0.000000000000001, 'p':0.000000000001, 'n':0.000000001, 'u':0.000001}
        def text_to_num2(text):
            if text[-1] in d2:
                num, magnitude = text[:-1], text[-1]
                return float(num) * d2[magnitude]
            else:
                return Decimal(text)
        new_c2value = text_to_num2(self.nimbleData['c2value'])
        
        # RVALUE Value to Position
        # If the value is less than or equal to 0 then we use minpos  
        def val_to_pos_rvalue(value):
            minpos = 1
            maxpos = 10000
            minval = math.log(10)
            maxval = math.log(10000000)
            scale = (maxval - minval) / (maxpos - minpos)
            global rposition
            rposition = minpos + (math.log(value) - minval) / scale
            return (rposition)
        val_to_pos_rvalue(new_rvalue)
        
        # C2 Value to Position
        def val_to_pos_c2value(value):
            minpos = 1
            maxpos = 10000
            minval = math.log(1e-13)
            maxval = math.log(0.000001)
            scale = (maxval - minval) / (maxpos - minpos)
            global c2position
            c2position = minpos + (math.log(value) - minval) / scale
            return (c2position)
        val_to_pos_c2value(new_c2value)

        # Accept cookies
        WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.XPATH, "//body/div[@id='base-container']/div[@id='noise-spinner']/div[1]")))
        time.sleep(7)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="cookie-consent-container"]/div/div/div[2]/div[1]/a'))).click()
        time.sleep(1)
        
        # Click on already dragged Amplifier
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/div[@id='build-signal-chain-tab-content']/div[@id='adi-signal-chain-row']/div[@id='analog-signal-chain-group']/div[@id='signal-chain-drop-area']/table[1]/tr[1]/td[1]/div[1]/div[2]/div[2]/*[1]"))).click()

        # Amp Configuration
        time.sleep(1)
        driver.execute_script(f"document.querySelector('#amp-gain-2').value={self.nimbleData['gain']}")
        
        # Select specific AMP
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#text6747-2-6 > tspan.schematic-edit-icon.schematic-part-edit-selection-link.schematic-edit-selection-link"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-0"))).send_keys(self.nimbleData['device'])
        time.sleep(1)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#device-table > div.slick-pane.slick-pane-top.slick-pane-left > div.slick-viewport.slick-viewport-top.slick-viewport-left > div > div"))).click()
        time.sleep(1)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#partSelectModal2 > div.modal.fade.in.show > div > div > form > div > button.btn.btn-primary"))).click()        
        time.sleep(5)

        # Resitance slider value
        self.scrollToRValue(rposition, driver)
        
        # C2 slider value
        self.scrollToC2Value(c2position, driver)

        # Use this Amplifier button
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#config-signal-chain-item-modal > div.modal.fade.in.show > div > div > form > div > button.btn.btn-primary"))).click()
        
        #Select Next Steps tab
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
        time.sleep(1)

        device = self.nimbleData['device']
        downloads_path = self.nimbleData['downloads_path']
        gain = self.nimbleData['gain']
        #current_date = self.nimbleData['current_date']
        current_date = datetime.today().strftime(("%B %d, %Y"))
        print("date: ", current_date)

        l = driver.current_url
        device_url = device + 'URL G' + gain + '.txt'
        with open(device_url, 'w') as f:
            f.write(l)
        print(l)
        time.sleep(1)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.tab-content:nth-child(2) div.download-area div.download-individual-buttons div.download-button-row:nth-child(1) button.btn.btn-primary:nth-child(1) > span:nth-child(1)"))).click()
        time.sleep(1)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.tab-content:nth-child(2) div.download-area div.download-individual-buttons div.download-button-row:nth-child(1) button.btn.btn-primary:nth-child(2) > span:nth-child(1)"))).click()

        time.sleep(3)
        #project_path = os.getcwd()
        project_path = self.nimbleData['project_location']

        if not os.path.exists(project_path + '\\' + device):
            os.makedirs(project_path + '\\' + device)
        dir_list = os.listdir()
        #print(dir_list)
        #print(project_path + device)

        ltspice_download_path = downloads_path + 'LTSpice ' + current_date + '.zip'
        shutil.move(ltspice_download_path, project_path + '/' + device)
        nimble_download_path = downloads_path + 'Raw Data Export - ' + current_date + '.zip'
        shutil.move(nimble_download_path, project_path + '/' + device)
        device_download_path = device + 'URL G' + gain + '.txt'
        shutil.move(device_download_path, project_path + '/' + device)
        time.sleep(1)

        downloaded_nimble_path = project_path + '\\' + device + '\\' + 'Raw Data Export - ' + current_date + '.zip'
        new_nimble_name = project_path + '\\' + device + '\\' + 'Nimble - ' + device + ' G' + gain + '.zip'
        downloaded_ltspice_path = project_path + '\\' + device + '\\' + 'LTspice ' + current_date + '.zip'
        new_ltspice_name = project_path + '\\' + device + '\\' + 'LTspice - ' + device + ' G' + gain + '.zip'
        os.rename(downloaded_nimble_path, new_nimble_name)
        os.rename(downloaded_ltspice_path, new_ltspice_name)

        with open(r'paths.json') as d:
            paths = json.load(d)['Nimble'][0]

        #Path Variables
        project_location = paths['project_location']
        device = paths['device']
        gain = paths['gain']

        # Unzip the two files from Nimble
        unzip_path_ltspice = project_location + '\\' + device + '\\' + 'LTspice - ' + device + ' G' + gain + '.zip'
        with zipfile.ZipFile(unzip_path_ltspice) as zip_ref:
            new_path = project_location + '\\' + device
            zip_ref.extractall(new_path)
        unzip_path_nimble = project_location + '\\' + device + '\\' + 'Nimble - ' + device + ' G' + gain + '.zip'
        with zipfile.ZipFile(unzip_path_nimble) as zip_ref:
            zip_ref.extractall(new_path)

        # Move Transfer Function csv to device folder
        time.sleep(1)
        raw_data = project_location + '\\' + device + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Transfer Function.csv'
        destination = project_location + '\\' + device
        shutil.move(raw_data, destination)
        time.sleep(3)

        # Path Variables
        project_loc = paths['project_location']
        device = paths['device']
        gain = paths['gain']

        # Open and save text tile from LTSpice
        os.startfile(project_loc + '\\' + device + '\\' + 'AC_Simulation.asc')
        pywinauto.keyboard.send_keys("%{S}")
        pywinauto.keyboard.send_keys("{R}")
        pywinauto.keyboard.send_keys("%{V}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("{V}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("{DOWN}")
        pywinauto.keyboard.send_keys("{DOWN}")
        pywinauto.keyboard.send_keys("{DOWN}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("{UP}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("{ENTER}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("^{TAB}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("%{F}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("{E}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("{ENTER}")
        time.sleep(1)
        pywinauto.keyboard.send_keys("%{F4}")

        time.sleep(0.5)

        # deleting extra files
        extra_files_remove = project_loc + device
        zip_remove_nimble = extra_files_remove + '\\' + 'Nimble - ' + device + ' G' + gain + '.zip'
        zip_remove_ltspice = extra_files_remove + '\\' + 'LTspice - ' + device + ' G' + gain + '.zip'

        # Converting the Transfer Function .csv to .xlsx
        path_file = pd.read_csv(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Transfer Function.csv')
        path_file.to_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Transfer Function.xlsx', index=None, header=True)
        excel_path = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Transfer Function.xlsx'

        # Deleting the extra collumn
        file = openpyxl.load_workbook(excel_path)
        sheet_obj = file.active
        sheet_obj.delete_cols(3)
        time.sleep(0.25)
        sheet_obj.delete_cols(3)
        time.sleep(0.25)
        sheet_obj.delete_cols(3)
        time.sleep(0.25)
        sheet_obj.delete_cols(3)
        file.save(excel_path)

        # Path Variables
        text_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.txt'
        output_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.xlsx'
        gain_textfile = paths['project_location'] + '\\' + paths['device'] + '\\' + paths['device'] + ' G' + paths['gain'] + '.txt'

        # Getting the data from AC_Simulation.txt to a new AC_Simulation.xlsx file
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        with open(text_file, 'r') as data:
            reader = csv.reader(data, delimiter='\t')
            for row in reader:
                ws.append(row)
        wb.save(output_file)

        # This script splits the test from the excel to get the data we need
        df = pd.read_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.xlsx')
        df.rename(columns={'V(out)':'vout'}, inplace=True)
        df['vout'] = df['vout'].str.split('(', expand=True)[1]
        df['vout'] = df['vout'].str.split('d', expand=True)[0]
        df['vout'] = '=VALUE(' + df['vout'] + ')'
        df.to_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.xlsx')

        # Getting the data from AC_Simulation.xlsx to Transfer Function.xlsx in a new sheet
        path1 = output_file
        path2 = excel_path

        wb1 = openpyxl.load_workbook(filename=path1)
        ws1 = wb1.worksheets[0]
        wb2 = openpyxl.load_workbook(filename=path2)
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(path2)

        # Transfering data from the new sheet to the sheet where the graph will be
        xl = openpyxl.load_workbook(excel_path)
        sheet1 = xl['Sheet11']
        sheet2 = xl['Sheet1']

        columnA = []
        for i in range(1, 1000, 1):
            columnA.append(sheet1.cell(row=i, column=2).value)
        for i in range(1, 1000, 1):
            for i in range(1, 1000, 1):
                sheet2.cell(row=i, column=4).value = columnA[i - 1]

        columnB = []
        for i in range(1, 1000, 1):
            columnB.append(sheet1.cell(row=i, column=3).value)
        for i in range(1, 1000, 1):
            for i in range(1, 1000, 1):
                sheet2.cell(row=i, column=5).value = columnB[i - 1]

        if 'Sheet11' in xl.sheetnames:
            xl.remove(xl['Sheet11'])
        xl.save(excel_path)

        # Deleting the extra files
        extra_files_remove = paths['project_location'] + '\\' + paths['device']
        zip_remove_nimble = extra_files_remove + '\\' + 'Nimble - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
        zip_remove_ltspice = extra_files_remove + '\\' + 'LTspice - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
        os.remove(extra_files_remove + '\\' + 'AC_Simulation.asc')
        os.remove(extra_files_remove + '\\' + 'AC_Simulation.log')
        os.remove(extra_files_remove + '\\' + 'AC_Simulation.op.raw')
        os.remove(extra_files_remove + '\\' + 'AC_Simulation.raw')
        os.remove(extra_files_remove + '\\' + 'AC_Simulation.txt')
        os.remove(extra_files_remove + '\\' + 'AC_Simulation.xlsx')
        os.remove(extra_files_remove + '\\' + 'Amplifier - Transfer Function.csv')
        os.remove(extra_files_remove + '\\' + 'Noise_Simulation.asc')
        os.remove(extra_files_remove + '\\' + 'Transient_Simulation.asc')
        os.remove(zip_remove_nimble)
        os.remove(zip_remove_ltspice)
        shutil.rmtree(extra_files_remove + '\\' + 'Raw Data')

        with open(r'paths.json') as d:
            paths = json.load(d)['Nimble'][0]

        workbook_path = (paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Transfer Function.xlsx')
        #"file_location": "C:\\Users\\psuatean\\OneDrive - ENDAVA\\Documents\\Python\\NimbleTest\\LTC6228\\Amplifier - Transfer Function.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook['Sheet1']
        sheet.title = ("G " + gain)

        sheet.cell(row=1, column=1).value = "Nimble - Freq."
        sheet.cell(row=1, column=2).value = "Nimble - Mag."
        sheet.cell(row=1, column=4).value = "LTSpice - Freq."
        sheet.cell(row=1, column=5).value = "LTSpice - Mag."

        for i in range(1,21):
            sheet.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

        x_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1010)
        y_nimble = Reference(sheet, min_col=1, min_row=2, max_row=1010)
        x_ltspice = Reference(sheet, min_col=5, min_row=2, max_row=1010)
        y_ltspice = Reference(sheet, min_col=4, min_row=2, max_row=1010)

        series_voltage = Series(x_nimble, y_nimble,title_from_data=False, title="Nimble")
        series_freq = Series(x_ltspice, y_ltspice,title_from_data=False, title="LTspice")

        # Chart type
        chart = ScatterChart()
        chart.series.append(series_voltage)
        chart.series.append(series_freq)

        chart.x_axis.scaling.logBase = 10
        chart.y_axis.number_format = '0.00E+00'
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.tickLblSkip = 3

        chart.x_axis.scaling.min = paths['x_axis_min']
        chart.y_axis.scaling.min = paths['y_axis_min']
        chart.x_axis.scaling.max = paths['x_axis_max']
        chart.y_axis.scaling.max = paths['y_axis_max']
        chart.x_axis.tickLblPos = "low"

        chart.title = None
        chart.x_axis.title = 'Frequency (Hz)'
        chart.y_axis.title = 'Magnitude (dB)'
        chart.legend.position = 'r'

        sheet.add_chart(chart, 'J2')
        workbook.save(workbook_path)

        rename_noise_excel = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Transfer Function.xlsx'
        new_noise_excel = paths['project_location'] + '\\' + paths['device'] + '\\' + paths['device'] + '.xlsx'
        os.rename(rename_noise_excel, new_noise_excel)
        print("Test Complete")

    @staticmethod
    def scrollToRValue(value: int, driver):
        driver.execute_script(f"document.querySelector('#rscale-slider').value = {rposition}; document.querySelector('#rscale-slider').dispatchEvent(new Event('input'));")
         
    @staticmethod
    def scrollToC2Value(value: int, driver):
        driver.execute_script(f"document.querySelector('#c2-slider').value = {c2position}; document.querySelector('#c2-slider').dispatchEvent(new Event('input'));")    

    def tearDown(self):
        #self.driver.quit()
        pass

if __name__ == '__main__':
    unittest.main()
